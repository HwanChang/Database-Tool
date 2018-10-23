# -*- coding: utf-8 -*-
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from openpyxl.styles import Border, Side, Font
import openpyxl, cx_Oracle, datetime, collections, threading
import os
os.putenv("NLS_LANG", "KOREAN_KOREA.KO16KSC5601")

class Oracle_Tibero:
	def __init__(self, info, textB):
		self.info = info
		self.textB = textB
		self.info['Progress'].start()
	# Functions by button type.
		if self.info['Type'] == 'ED' or self.info['Type'] == 'ES':
			if self.info['Type'] == 'ED' and self.info['Drop'] == 1:
				self.dropT, self.dropS = list(), list()
				self.info['Cursor'].execute('SELECT TABLE_NAME FROM tabs')
				self.tab = self.info['Cursor'].fetchall()
				self.info['Cursor'].execute('SELECT SEQUENCE_NAME FROM user_sequences')
				self.seq = self.info['Cursor'].fetchall()
			self.sendSQL = collections.OrderedDict()
			self.realList = list()
			self.excel_document = openpyxl.load_workbook(self.info['Path'])
			self.sheetList = self.excel_document.sheetnames
			if self.info['Sheet'] == 'all':
				for sheetnameList in self.sheetList:
					if sheetnameList[0] == '#':
						continue
					self.realList.append(sheetnameList)
				for real in self.realList:
					self.ED_ESFunction(real)
			else:
				self.realList.append(self.info['Sheet'])
				self.ED_ESFunction(self.info['Sheet'])
		elif self.info['Type'] == 'DE':
			self.DB_ExcelFunction()
		elif self.info['Type'] == 'DS':
			self.DB_SQLFunction()

# Main function about excel -> DB Scheme, Excel -> SQL File.
	def ED_ESFunction(self, name):
		self.name = name
		self.sheet = self.excel_document[name]
		self.sendSQL[name] = list()
		self.tables = collections.OrderedDict()
		self.commentsTables = collections.OrderedDict()
		count = 0
		for row in list(self.sheet.rows)[1:]:
			if row[0].value == '#':
				continue
			if row[1].value is None:
				count += 1
				if count > 5:
					self.tables[tableName[0]] = tableRows
					self.commentsTables[tableName[0]] = tableName[1]
					break
				continue
			if row[3].value is None:
				count = 0
				try:
					self.tables[tableName[0]] = tableRows
					self.commentsTables[tableName[0]] = tableName[1]
				except UnboundLocalError:
					pass
				tableRows = list()
				tableName = [str(row[1].value), row[2].value]
				continue
			length = str()
			korName = str()
			if row[2].value is not None:
				korName = row[2].value
			if row[4].value is not None:
				length = str(row[4].value)
			tableRows.append([str(row[1].value), korName, str(row[3].value), length, str(row[5].value)])
			if row == list(self.sheet.rows)[-1]:
				self.tables[tableName[0]] = tableRows
				self.commentsTables[tableName[0]] = tableName[1]
		commentsColumns = collections.OrderedDict()
		for tblName, tblContents in self.tables.items():
			SQL = ''
			constraintName = list()
			commentsColumns[tblName] = list()
			notNull = False
			for row in tblContents:
				if row == tblContents[0]:
					if row[4] == 'Y':
						if row[2] == 'string':
							SQL += '\t' + row[0] + ' varchar2(' + row[3] + ') NOT NULL'
						elif row[2] == 'char':
							SQL += '\t' + row[0] + ' ' + row[2] + '(' + row[3] + ') NOT NULL'
						elif 'number' in row[2]:
							SQL += '\t' + row[0] + ' ' + row[2] + ' NOT NULL'
						constraintName.append(row[0])
						notNull = True
						commentsColumns[tblName].append([row[0], row[1]])
					else:
						if row[2] == 'int':
							SQL += '\t' + row[0] + ' number PRIMARY KEY NOT NULL'
						elif row[2] == 'string':
							SQL += '\t' + row[0] + ' varchar2(' + row[3] + ')'
						elif row[2] == 'char':
							SQL += '\t' + row[0] + ' ' + row[2] + '(' + row[3] + ')'
						elif row[2] == 'text':
							SQL += '\t' + row[0] + ' clob'
						elif 'number' in row[2]:
							SQL += '\t' + row[0] + ' ' + row[2] + ''
						commentsColumns[tblName].append([row[0], row[1]])
				else:
					if row[4] == 'Y':
						if row[2] == 'int':
							SQL += ', \n\t' + row[0] + ' number NOT NULL'
						elif row[2] == 'string':
							SQL += ', \n\t' + row[0] + ' varchar2(' + row[3] + ') NOT NULL'
						elif row[2] == 'char':
							SQL += ', \n\t' + row[0] + ' ' + row[2] + '(' + row[3] + ') NOT NULL'
						elif 'number' in row[2]:
							SQL += ', \n\t' + row[0] + ' ' + row[2] + ' NOT NULL'
						commentsColumns[tblName].append([row[0], row[1]])
						constraintName.append(row[0])
						notNull = True
					else:
						if row[2] == 'int':
							SQL += ', \n\t' + row[0] + ' number'
						elif row[2] == 'string':
							SQL += ', \n\t' + row[0] + ' varchar2(' + row[3] + ')'
						elif row[2] == 'char':
							SQL += ', \n\t' + row[0] + ' ' + row[2] + '(' + row[3] + ')'
						elif row[2] == 'text':
							SQL += ', \n\t' + row[0] + ' clob'
						elif 'number' in row[2]:
							SQL += ', \n\t' + row[0] + ' ' + row[2] + ''
						commentsColumns[tblName].append([row[0], row[1]])
			if notNull:
				constList = ''
				for index_C, cont in enumerate(constraintName):
					if index_C is not len(constraintName)-1:
						constList += cont + ', '
					else:
						constList += cont
				SQL += ', \n\n\tCONSTRAINT UK_' + tblName.split('_')[1] + ' UNIQUE(' + constList + ')'
			comments = list()
			for comment in commentsColumns[tblName]:
				if comment == commentsColumns[tblName][0]:
					comments.append('COMMENT ON TABLE ' + tblName + " IS '" + self.commentsTables[tblName] + "'")
				comments.append('COMMENT ON COLUMN ' + tblName + '.' + comment[0] + " IS '" + comment[1] + "'")
			self.sendSQL[name].append(['CREATE TABLE ' + tblName + ' \n(\n' + SQL + '\n)\n', '\n\nCREATE SEQUENCE SEQ_' + tblName.split('_')[1] + ' \nINCREMENT BY 1\nSTART WITH 1\nNOMAXVALUE\nNOCYCLE\nNOCACHE\n', comments])
	# Excel -> DB Scheme function.
		if self.info['Type'] == 'ED':
			if self.info['Drop'] == 1:
				for send in self.sendSQL[self.name]:
					for tName in self.tab:
						if tName[0] == str(send[0].split(' ')[2]).upper():
							self.dropT.append(tName[0])
							break
				for send in self.sendSQL[self.name]:
					for sName in self.seq:
						if sName[0] == str(send[1].split(' ')[2]).upper():
							self.dropS.append(sName[0])
							break
			for send in self.sendSQL[self.name]:
				if self.info['Drop'] == 1:
					if str(send[0].split(' ')[2]).upper() in self.dropT:
						self.info['Cursor'].execute('DROP TABLE ' + str(send[0].split(' ')[2]))
					self.info['Cursor'].execute(send[0])
					if str(send[1].split(' ')[2]).upper() in self.dropS:
						self.info['Cursor'].execute('DROP SEQUENCE ' + str(send[1].split(' ')[2]))
					self.info['Cursor'].execute(send[1])
				else:
					self.info['Cursor'].execute(send[0])
					self.info['Cursor'].execute(send[1])
				for sql in send[2]:
					self.info['Cursor'].execute(sql)
			if name == self.realList[-1]:
				self.info['Progress'].stop()
				self.textB.delete(1.0, END)
				self.textB.insert(1.0, 'Excel -> DB Scheme Complete!')
				f = open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'a')
				f.write(str('%s-%s-%s %s:%s:%s' %(datetime.datetime.now().year, datetime.datetime.now().month, datetime.datetime.now().day, datetime.datetime.now().hour, datetime.datetime.now().minute, datetime.datetime.now().second)) + '\tExcel -> DB Scheme Function.\t\t[ ' + str(self.info['Path'].encode('euc-kr')) + ' ]\n')
				f.close()
	# Excel -> SQL File Function.
		elif self.info['Type'] == 'ES' and name == self.realList[-1]:
			self.Excel_SQLFunction()

	def Excel_SQLFunction(self):
		try:
			f = open(self.info['ESSave'], 'w')
			if self.info['Sheet'] == 'all':
				for realSheetList in self.realList:
					for send in self.sendSQL[realSheetList]:
						commStr = str()
						for comm in send[2]:
							commStr += comm + ';\n\n'
						f.write(send[0] + ';' + send[1] + ';\n\n' + commStr)
			else:
				for send in self.sendSQL[self.name]:
					commStr = str()
					for comm in send[2]:
						commStr += comm + ';\n\n'
					f.write(send[0] + ';' + send[1] + ';\n\n' + commStr)
			f.close()
			self.info['Progress'].stop()
			self.textB.insert(1.0, 'Excel -> SQL File Complete!\n\n')
			self.textB.insert(END, self.info['ESSave'])
			f = open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'a')
			f.write(str('%s-%s-%s %s:%s:%s' %(datetime.datetime.now().year, datetime.datetime.now().month, datetime.datetime.now().day, datetime.datetime.now().hour, datetime.datetime.now().minute, datetime.datetime.now().second)) + '\tExcel -> SQL File Function.\t\t\t[ ' + str(self.info['Path'].encode('euc-kr')) + ' -> ' + self.info['ESSave'] + ' ]\n')
			f.close()
		except IOError:
			self.info['Progress'].stop()
			messagebox.showwarning('Warning','Please select a SQL file to save.')

	def DB_ExcelFunction(self):
		self.info['Progress'].start()
		try:
			wb = openpyxl.load_workbook('C:\\Users\\Secuve\\Desktop\\Database Tool\\history file\\history.xlsx')
			sheetmkNew = wb.create_sheet()
			sheetmkNew.title = self.info['DESheet']
		except ValueError:
			self.info['Progress'].stop()
			messagebox.showwarning('Warning','Please fill out the Sheet name.')

		sheetmkNew.column_dimensions['A'].width = 5
		sheetmkNew.column_dimensions['B'].width = 30
		sheetmkNew.column_dimensions['C'].width = 30
		sheetmkNew.column_dimensions['D'].width = 15
		sheetmkNew.column_dimensions['E'].width = 10
		sheetmkNew.column_dimensions['F'].width = 7
		sheetmkNew.column_dimensions['G'].width = 10
		sheetmkNew.column_dimensions['H'].width = 30
		sheetmkNew.column_dimensions['I'].width = 60

		sheetmkNew.cell(row=1, column=1).value = u'주석'
		sheetmkNew.cell(row=1, column=2).value = u'필드명'
		sheetmkNew.cell(row=1, column=3).value = u'필드명(한글)'
		sheetmkNew.cell(row=1, column=4).value = u'데이터타입'
		sheetmkNew.cell(row=1, column=5).value = u'길이'
		sheetmkNew.cell(row=1, column=6).value = u'필수'
		sheetmkNew.cell(row=1, column=7).value = u'유효길이'
		sheetmkNew.cell(row=1, column=8).value = u'샘플데이터'
		sheetmkNew.cell(row=1, column=9).value = u'설명'
		sheetmkNew.cell(row=2, column=6).value = 'Y/N'

		fontBold = Font(bold = True)
		column_border_L = Border(left=Side(style='thick'))
		column_border_R = Border(right=Side(style='thick'))
		row_border_T = Border(top=Side(style='thick'))
		row_border_B = Border(bottom=Side(style='thick'))

		self.textB.delete(1.0, END)

		lineCnt = 3
		for index in self.info['DEListBox'].curselection():
			tableName = str(self.info['DEListBox'].get(index)[0])
			self.info['Cursor'].execute("SELECT U.COLUMN_NAME, U.DATA_TYPE, U.DATA_LENGTH, A.COMMENTS FROM USER_TAB_COLUMNS U, ALL_COL_COMMENTS A WHERE U.COLUMN_NAME = A.COLUMN_NAME AND U.TABLE_NAME = '" + tableName + "' AND A.TABLE_NAME = '" + tableName + "'")
			sqlList = self.info['Cursor'].fetchall()
			self.info['Cursor'].execute("SELECT S.CONSTRAINT_TYPE, C.COLUMN_NAME FROM USER_CONS_COLUMNS C INNER JOIN USER_CONSTRAINTS S ON C.CONSTRAINT_NAME = S.CONSTRAINT_NAME AND (S.CONSTRAINT_TYPE = 'P' OR S.CONSTRAINT_TYPE = 'U') WHERE C.TABLE_NAME = '" + tableName + "' ORDER BY 1")
			constraintList = self.info['Cursor'].fetchall()
			self.info['Cursor'].execute("SELECT COMMENTS FROM USER_TAB_COMMENTS WHERE TABLE_NAME = '" + tableName + "'")
			tableComment = self.info['Cursor'].fetchall()
			rowList = list()
			sheetmkNew.cell(row=lineCnt, column=2).value = tableName
			sheetmkNew.cell(row=lineCnt, column=3).value = tableComment[0][0]
			sheetmkNew['B' + str(lineCnt)].font = fontBold
			sheetmkNew['C' + str(lineCnt)].font = fontBold
			lineCnt += 1
			for sqllist in sqlList:
				rowList.append([str(sqllist[0]), str(sqllist[1]), str(int(sqllist[2])), str(sqllist[3])])
			for row in rowList:
				if row == rowList[0]:
					for rowC in sheetmkNew['B' + str(lineCnt) + ':I' + str(lineCnt)]:
						for cell in rowC:
							cell.border = cell.border + row_border_T
				if row == rowList[-1]:
					for rowC in sheetmkNew['B' + str(lineCnt) + ':I' + str(lineCnt)]:
						for cell in rowC:
							cell.border = cell.border + row_border_B
				sheetmkNew['B' + str(lineCnt)].border = sheetmkNew['B' + str(lineCnt)].border + column_border_L
				sheetmkNew['I' + str(lineCnt)].border = sheetmkNew['I' + str(lineCnt)].border + column_border_R
				check = True
				lowerRow1 = row[1].lower()
				for const in constraintList:
					if const[1] == row[0]:
						if const[0] == 'P':
							sheetmkNew.cell(row=lineCnt, column=2).value = row[0]
							sheetmkNew.cell(row=lineCnt, column=3).value = row[3]
							sheetmkNew.cell(row=lineCnt, column=4).value = 'int'
							sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
							sheetmkNew.cell(row=lineCnt, column=6).value = ''
							lineCnt += 1
							check = False
							break
						elif const[0] == 'U':
							sheetmkNew.cell(row=lineCnt, column=2).value = row[0]
							sheetmkNew.cell(row=lineCnt, column=3).value = row[3]
							if lowerRow1 == 'varchar2':
								sheetmkNew.cell(row=lineCnt, column=4).value = 'string'
							elif lowerRow1 == 'clob':
								sheetmkNew.cell(row=lineCnt, column=4).value = 'text'
							else:
								sheetmkNew.cell(row=lineCnt, column=4).value = lowerRow1
							sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
							sheetmkNew.cell(row=lineCnt, column=6).value = 'Y'
							lineCnt += 1
							check = False
							break
				if check:
					sheetmkNew.cell(row=lineCnt, column=2).value = row[0]
					sheetmkNew.cell(row=lineCnt, column=3).value = row[3]
					if lowerRow1 == 'varchar2':
						sheetmkNew.cell(row=lineCnt, column=4).value = 'string'
					elif lowerRow1 == 'clob':
						sheetmkNew.cell(row=lineCnt, column=4).value = 'text'
					else:
						sheetmkNew.cell(row=lineCnt, column=4).value = lowerRow1
					sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
					sheetmkNew.cell(row=lineCnt, column=6).value = ''
					lineCnt += 1
			lineCnt += 1
		try:
			wb.save(self.info['DEPath'])
			self.info['Progress'].stop()
			self.info['Window'].destroy()
			self.textB.insert(1.0, 'DB Scheme -> Excel Complete!\n\n')
			self.textB.insert(END, self.info['DEPath'])
			f = open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'a')
			f.write(str('%s-%s-%s %s:%s:%s' %(datetime.datetime.now().year, datetime.datetime.now().month, datetime.datetime.now().day, datetime.datetime.now().hour, datetime.datetime.now().minute, datetime.datetime.now().second)) + '\tDB Scheme -> Excel File Function.\t[ ' + self.info['DEPath'] + ' ]\n')
			f.close()
		except IOError:
			self.info['Progress'].stop()
			messagebox.showwarning('Warning','Please select a Excel file to save.')

	def DB_SQLFunction(self):
		self.info['Progress'].start()
		try:
			writeSQLsentence = ''
			for index in self.info['DSListBox'].curselection():
				tableName = str(self.info['DSListBox'].get(index)[0])
				checkNull = False
				self.info['Cursor'].execute("SELECT U.COLUMN_NAME, U.DATA_TYPE, U.DATA_LENGTH, A.COMMENTS FROM USER_TAB_COLUMNS U, ALL_COL_COMMENTS A WHERE U.COLUMN_NAME = A.COLUMN_NAME AND U.TABLE_NAME = '" + tableName + "' AND A.TABLE_NAME = '" + tableName + "'")
				sqlList = self.info['Cursor'].fetchall()
				self.info['Cursor'].execute("SELECT S.CONSTRAINT_TYPE, C.COLUMN_NAME FROM USER_CONS_COLUMNS C INNER JOIN USER_CONSTRAINTS S ON C.CONSTRAINT_NAME = S.CONSTRAINT_NAME AND (S.CONSTRAINT_TYPE = 'P' OR S.CONSTRAINT_TYPE = 'U' OR S.CONSTRAINT_TYPE = 'C') WHERE C.TABLE_NAME = '" + tableName + "' ORDER BY 1")
				constraintList = self.info['Cursor'].fetchall()
				self.info['Cursor'].execute("SELECT COMMENTS FROM USER_TAB_COMMENTS WHERE TABLE_NAME = '" + tableName + "'")
				tableComment = self.info['Cursor'].fetchall()
				SQLsentence = 'CREATE TABLE ' + tableName + ' \n(\n'
				rowList = list()
				constraintName = list()
				constCheck = list()
				for sqllist in sqlList:
					rowList.append([str(sqllist[0]), str(sqllist[1]), str(int(sqllist[2])), str(sqllist[3])])
				for const in constraintList:
					if const[0] == 'P' or const[0] == 'U':
						constCheck.append(const[1])
				for row in rowList:
					flag = True
					if row == rowList[-1]:
						for const in constraintList:
							if const[1] == row[0]:
								if const[0] == 'P':
									SQLsentence += '\t' + row[0] + ' ' + row[1] + ' PRIMARY KEY NOT NULL'
									flag = False
									break
								elif const[0] == 'U':
									SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ') NOT NULL'
									flag = False
									constraintName.append(row[0])
									checkNull = True
									break
								elif const[0] == 'C':
									if const[0] not in constCheck:
										SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ') NOT NULL'
										flag = False
										break
							if const == constraintList[-1]:
								flag = True
						if flag:
							SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ')'
					else:
						for const in constraintList:
							if const[1] == row[0]:
								if const[0] == 'P':
									SQLsentence += '\t' + row[0] + ' ' + row[1] + ' PRIMARY KEY NOT NULL, \n'
									flag = False
									break
								elif const[0] == 'U':
									SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ') NOT NULL, \n'
									flag = False
									constraintName.append(row[0])
									checkNull = True
									break
								elif const[0] == 'C':
									if const[0] not in constCheck:
										SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ') NOT NULL, \n'
										flag = False
										break
							if const == constraintList[-1]:
								flag = True
						if flag:
							SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + '), \n'
				if checkNull:
					constList = ''
					for const in constraintName:
						if const != constraintName[-1]:
							constList += const + ', '
						else:
							constList += const
					SQLsentence += ',\n\n\tCONSTRAINT UK_' + str(self.info['DSListBox'].get(index)[0]).split('_')[1] + ' UNIQUE(' + constList + ')'
				SQLsentence = SQLsentence + '\n)\n;\n\nCREATE SEQUENCE SEQ_' + tableName.split('_')[1] + ' \nINCREMENT BY 1\nSTART WITH 1\nNOMAXVALUE\nNOCYCLE\nNOCACHE\n;\n\n'
				cCStr = str()
				for cC in rowList:
					cCStr += 'COMMENT ON COLUMN ' + tableName + '.' + cC[0] + " IS '" + cC[3] + "';\n\n"
				writeSQLsentence += SQLsentence + 'COMMENT ON TABLE ' + tableName + " IS '" + tableComment[0][0] + "';\n\n" + cCStr
			f = open(self.info['DSPath'], 'w')
			f.write(writeSQLsentence)
			f.close()
			self.info['Progress'].stop()
			self.info['Window'].destroy()
			self.textB.insert(1.0, 'DB Scheme -> SQL File Complete!\n\n')
			self.textB.insert(END, self.info['DSPath'])
			f = open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'a')
			f.write(str('%s-%s-%s %s:%s:%s' %(datetime.datetime.now().year, datetime.datetime.now().month, datetime.datetime.now().day, datetime.datetime.now().hour, datetime.datetime.now().minute, datetime.datetime.now().second)) + '\tDB Scheme -> SQL File Function.\t\t[ ' + self.info['DSPath'] + ' ]\n')
			f.close()
		except IOError:
			self.info['Progress'].stop()
			messagebox.showwarning('Warning','Please select a SQL file to save.')
