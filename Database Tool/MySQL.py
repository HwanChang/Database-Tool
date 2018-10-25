# -*- coding: utf-8 -*-
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from openpyxl.styles import Border, Side, Font
import openpyxl, pymysql, datetime, collections, threading, MainFrame

class MySQL:
	def __init__(self, info, textB):
		self.info = info
		self.textB = textB
		self.datetime = datetime.datetime.now()
	# Functions by button type.
		if self.info['Type'] == 'ED' or self.info['Type'] == 'ES':
			if self.info['Type'] == 'ED' and self.info['Drop'] == 1:
				self.dropT =  list()
				self.info['Cursor'].execute('SHOW tables')
				self.tab = self.info['Cursor'].fetchall()
			self.sendSQL = collections.OrderedDict()
			self.realList = list()
			self.excel_document = openpyxl.load_workbook(self.info['Path'])
			self.sheetList = self.excel_document.sheetnames
			count = 1
			if self.info['Sheet'] == 'all':
				for sheetnameList in self.sheetList:
					if sheetnameList[0] == '#':
						continue
					self.realList.append(sheetnameList)
				self.info['Progress'].config(maximum=len(self.realList))
				for real in self.realList:
					self.ED_ESFunction(real)
					self.info['Progress'].config(value=count)
					self.info['Percent'].config(text=str(count)+' / '+str(len(self.realList)) + ' Sheets')
					count += 1
			else:
				self.realList.append(self.info['Sheet'])
				self.info['Progress'].config(maximum=len(self.realList))
				self.ED_ESFunction(self.info['Sheet'])
				self.info['Progress'].config(value=count)
				self.info['Percent'].config(text=str(count)+' / '+str(len(self.realList)) + ' Sheets')
				count += 1
		elif self.info['Type'] == 'DE':
			self.DB_ExcelFunction()
		elif self.info['Type'] == 'DS':
			self.DB_SQLFunction()
# Main function about excel -> DB Scheme, Excel -> SQL File.
	def ED_ESFunction(self, name):
		self.name = name
		self.sheet = self.excel_document[name]
		self.sendSQL[name] = list()
		self.tables = collections.OrderedDict()
		self.commentsTables = collections.OrderedDict()
		count = 0
		for row in list(self.sheet.rows)[1:]:
			if row[0].value == '#':
				continue
			if row[1].value is None:
				count += 1
				if count > 5:
					self.tables[tableName[0]] = tableRows
					self.commentsTables[tableName[0]] = tableName[1]
					break
				continue
			if row[3].value is None:
				count = 0
				try:
					self.tables[tableName[0]] = tableRows
					self.commentsTables[tableName[0]] = tableName[1]
				except UnboundLocalError:
					pass
				tableRows = list()
				tableName = [str(row[1].value), row[2].value]
				continue
			length = str()
			korName = str()
			if row[2].value is not None:
				korName = row[2].value
			if row[4].value is not None:
				length = str(row[4].value)
			tableRows.append([str(row[1].value), korName, str(row[3].value), length, str(row[5].value)])
			if row == list(self.sheet.rows)[-1]:
				self.tables[tableName[0]] = tableRows
				self.commentsTables[tableName[0]] = tableName[1]
		for tblName, tblContents in self.tables.items():
			SQL = ''
			constraintName = list()
			notNull = False
			for row in tblContents:
				if row == tblContents[0]:
					if row[4] == 'Y':
						if row[2] == 'string':
							SQL += '\t' + row[0] + ' varchar(' + row[3] + ') NOT NULL'
						elif row[2] == 'char':
							SQL += '\t' + row[0] + ' ' + row[2] + '(' + row[3] + ') NOT NULL'
						elif 'number' in row[2]:
							SQL += '\t' + row[0] + ' ' + row[2].replace('number', 'numeric') + ' NOT NULL'
						SQL += " COMMENT '" + row[1] + "'"
						constraintName.append(row[0])
						notNull = True
					else:
						if row[2] == 'int':
							SQL += '\t' + row[0] + ' int PRIMARY KEY NOT NULL'
						elif row[2] == 'string':
							SQL += '\t' + row[0] + ' varchar(' + row[3] + ')'
						elif row[2] == 'char':
							SQL += '\t' + row[0] + ' ' + row[2] + '(' + row[3] + ')'
						elif row[2] == 'text':
							SQL += '\t' + row[0] + ' LONGTEXT'
						elif 'number' in row[2]:
							SQL += '\t' + row[0] + ' ' + row[2].replace('number', 'numeric') + ''
						SQL += " COMMENT '" + row[1] + "'"
				else:
					if row[4] == 'Y':
						if row[2] == 'int':
							SQL += ', \n\t' + row[0] + ' number NOT NULL'
						elif row[2] == 'string':
							SQL += ', \n\t' + row[0] + ' varchar(' + row[3] + ') NOT NULL'
						elif row[2] == 'char':
							SQL += ', \n\t' + row[0] + ' ' + row[2] + '(' + row[3] + ') NOT NULL'
						elif 'number' in row[2]:
							SQL += ', \n\t' + row[0] + ' ' + row[2].replace('number', 'numeric') + ' NOT NULL'
						SQL += " COMMENT '" + row[1] + "'"
						constraintName.append(row[0])
						notNull = True
					else:
						if row[2] == 'int':
							SQL += ', \n\t' + row[0] + ' int'
						elif row[2] == 'string':
							SQL += ', \n\t' + row[0] + ' varchar(' + row[3] + ')'
						elif row[2] == 'char':
							SQL += ', \n\t' + row[0] + ' ' + row[2] + '(' + row[3] + ')'
						elif row[2] == 'text':
							SQL += ', \n\t' + row[0] + ' LONGTEXT'
						elif 'number' in row[2]:
							SQL += ', \n\t' + row[0] + ' ' + row[2].replace('number', 'numeric') + ''
						SQL += " COMMENT '" + row[1] + "'"
			if notNull:
				constList = ''
				for index_C, cont in enumerate(constraintName):
					if index_C is not len(constraintName)-1:
						constList += cont + ', '
					else:
						constList += cont
				SQL += ', \n\n\tCONSTRAINT UK_' + tblName.split('_')[1] + ' UNIQUE(' + constList + ')'
			self.sendSQL[name].append(['CREATE TABLE ' + tblName + ' \n(\n' + SQL + "\n) COMMENT '" + str(self.commentsTables[tblName]) + "'\n\nCOLLATE='utf8_bin'\nENGINE=InnoDB\n"])
	# Excel -> DB Scheme function.
		if self.info['Type'] == 'ED':
			try:
				if self.info['Drop'] == 1:
					for send in self.sendSQL[self.name]:
						for tName in self.tab:
							if tName[0] == str(send[0].split(' ')[2]):
								self.dropT.append(tName[0])
								break
				for send in self.sendSQL[self.name]:
					if self.info['Drop'] == 1:
						if str(send[0].split(' ')[2]) in self.dropT:
							self.info['Cursor'].execute('DROP TABLE ' + str(send[0].split(' ')[2]))
						self.info['Cursor'].execute(send[0])
					else:
						self.info['Cursor'].execute(send[0])
				if name == self.realList[-1]:
					self.info['Thread'].statusCheck = False
					self.info['Status'].join()
					self.textB.delete(1.0, END)
					self.textB.insert(1.0, 'Excel File -> DB Scheme Complete!\n\n')
					f = open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'a')
					f.write(self.datetime.strftime('[ %Y-%m-%d %H:%M:%S ]') + "%-40s" % '\t\tExcel File -> DB Scheme Function.' + "%-60s" % ('[ ' + self.info['Path']) + ' ]\n')
					f.close()
					with open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'r') as f:
						lines = f.readlines()
						if len(lines) > 20:
							for i in range(0, len(lines)-20):
								del lines[0]
						for line in lines:
							self.textB.insert(END, line)
					self.textB.config(state=DISABLED)
			except pymysql.InternalError as e:
				code, message = e.args
				if code == 1050:
					messagebox.showwarning('Warning', 'Please check the DB.\nThe table name is already used.')
	# Excel -> SQL File Function.
		elif self.info['Type'] == 'ES' and name == self.realList[-1]:
			self.Excel_SQLFunction()

	def Excel_SQLFunction(self):
		try:
			f = open(self.info['ESSave'], 'w')
			if self.info['Sheet'] == 'all':
				for realSheetList in self.realList:
					for send in self.sendSQL[realSheetList]:
						f.write(send[0] + ';\n\n')
			else:
				for send in self.sendSQL[self.name]:
					f.write(send[0] + ';\n\n')
			f.close()
			self.info['Thread'].stopFunction(False)
			self.info['Status'].join()
			self.textB.delete(1.0, END)
			self.textB.insert(1.0, 'Excel File -> SQL File Complete!\n\n')
			self.textB.insert(END, self.info['ESSave'] + '\n\n')
			f = open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'a')
			f.write(self.datetime.strftime('[ %Y-%m-%d %H:%M:%S ]') + "%-40s" % '\t\tExcel Fiel -> SQL File Function.' + "%-69s" % ('[ ' + self.info['ESSave']) + ' ]\n')
			f.close()
			with open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'r') as f:
				lines = f.readlines()
				if len(lines) > 20:
					for i in range(0, len(lines)-20):
						del lines[0]
				for line in lines:
					self.textB.insert(END, line)
			self.textB.config(state=DISABLED)
		except IOError:
			messagebox.showwarning('Warning','Please select a SQL file to save.')

	def DB_ExcelFunction(self):
		try:
			wb = openpyxl.load_workbook('C:\\Users\\Secuve\\Desktop\\Database Tool\\history file\\history.xlsx')
			sheetmkNew = wb.create_sheet()
			sheetmkNew.title = self.info['DESheet']
			sheetmkNew.column_dimensions['A'].width = 5
			sheetmkNew.column_dimensions['B'].width = 30
			sheetmkNew.column_dimensions['C'].width = 30
			sheetmkNew.column_dimensions['D'].width = 15
			sheetmkNew.column_dimensions['E'].width = 10
			sheetmkNew.column_dimensions['F'].width = 7
			sheetmkNew.column_dimensions['G'].width = 10
			sheetmkNew.column_dimensions['H'].width = 30
			sheetmkNew.column_dimensions['I'].width = 60

			sheetmkNew.cell(row=1, column=1).value = u'주석'
			sheetmkNew.cell(row=1, column=2).value = u'필드명'
			sheetmkNew.cell(row=1, column=3).value = u'필드명(한글)'
			sheetmkNew.cell(row=1, column=4).value = u'데이터타입'
			sheetmkNew.cell(row=1, column=5).value = u'길이'
			sheetmkNew.cell(row=1, column=6).value = u'필수'
			sheetmkNew.cell(row=1, column=7).value = u'유효길이'
			sheetmkNew.cell(row=1, column=8).value = u'샘플데이터'
			sheetmkNew.cell(row=1, column=9).value = u'설명'
			sheetmkNew.cell(row=2, column=6).value = 'Y/N'

			fontBold = Font(bold = True)
			column_border_L = Border(left=Side(style='thick'))
			column_border_R = Border(right=Side(style='thick'))
			row_border_T = Border(top=Side(style='thick'))
			row_border_B = Border(bottom=Side(style='thick'))

			self.info['Cursor'].execute("SELECT TABLE_NAME, TABLE_COMMENT FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='" + self.info['sid'] + "'")
			tableComment = self.info['Cursor'].fetchall()
			lineCnt = 3
			count = 1
			self.info['Progress'].config(maximum=len(self.info['DEListBox'].curselection()))
			for index in self.info['DEListBox'].curselection():
				tableName = str(self.info['DEListBox'].get(index)[0])
				self.info['Cursor'].execute("SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, COLUMN_COMMENT, IS_NULLABLE, COLUMN_TYPE FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA='" + self.info['sid'] + "' AND TABLE_NAME='" + tableName + "'")
				sqlList = self.info['Cursor'].fetchall()
				rowList = list()
				sheetmkNew.cell(row=lineCnt, column=2).value = tableName
				for comment in tableComment:
					if comment[0] == tableName:
						sheetmkNew.cell(row=lineCnt, column=3).value = comment[1]
				sheetmkNew['B' + str(lineCnt)].font = fontBold
				sheetmkNew['C' + str(lineCnt)].font = fontBold
				lineCnt += 1
				for sqllist in sqlList:
					rowList.append([str(sqllist[1]), str(sqllist[2]), str(sqllist[3]), str(sqllist[4]), str(sqllist[5]), str(sqllist[6])])
				for row in rowList:
					if row == rowList[0]:
						for rowC in sheetmkNew['B' + str(lineCnt) + ':I' + str(lineCnt)]:
							for cell in rowC:
								cell.border = cell.border + row_border_T
					if row == rowList[-1]:
						for rowC in sheetmkNew['B' + str(lineCnt) + ':I' + str(lineCnt)]:
							for cell in rowC:
								cell.border = cell.border + row_border_B
					sheetmkNew['B' + str(lineCnt)].border = sheetmkNew['B' + str(lineCnt)].border + column_border_L
					sheetmkNew['I' + str(lineCnt)].border = sheetmkNew['I' + str(lineCnt)].border + column_border_R
					check = True
					if row[4] == 'NO':
						if row[1] == 'int':
							sheetmkNew.cell(row=lineCnt, column=2).value = row[0]
							sheetmkNew.cell(row=lineCnt, column=3).value = row[3]
							sheetmkNew.cell(row=lineCnt, column=4).value = row[1]
							sheetmkNew.cell(row=lineCnt, column=5).value = ''
							sheetmkNew.cell(row=lineCnt, column=6).value = ''
							lineCnt += 1
						else:
							sheetmkNew.cell(row=lineCnt, column=2).value = row[0]
							sheetmkNew.cell(row=lineCnt, column=3).value = row[3]
							if row[1] == 'varchar':
								sheetmkNew.cell(row=lineCnt, column=4).value = 'string'
								sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
							elif row[1] == 'longtext':
								sheetmkNew.cell(row=lineCnt, column=4).value = 'text'
								sheetmkNew.cell(row=lineCnt, column=5).value = ''
							elif row[1] == 'decimal':
								sheetmkNew.cell(row=lineCnt, column=4).value = row[5].replace('decimal', 'number')
								sheetmkNew.cell(row=lineCnt, column=5).value = ''
							else:
								sheetmkNew.cell(row=lineCnt, column=4).value = row[1]
								sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
							sheetmkNew.cell(row=lineCnt, column=6).value = 'Y'
							lineCnt += 1
						check = False
					else:
						if row[1] == 'int':
							sheetmkNew.cell(row=lineCnt, column=2).value = row[0]
							sheetmkNew.cell(row=lineCnt, column=3).value = row[3]
							sheetmkNew.cell(row=lineCnt, column=4).value = row[1]
							sheetmkNew.cell(row=lineCnt, column=5).value = ''
							sheetmkNew.cell(row=lineCnt, column=6).value = ''
							lineCnt += 1
						else:
							sheetmkNew.cell(row=lineCnt, column=2).value = row[0]
							sheetmkNew.cell(row=lineCnt, column=3).value = row[3]
							if row[1] == 'varchar':
								sheetmkNew.cell(row=lineCnt, column=4).value = 'string'
								sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
							elif row[1] == 'longtext':
								sheetmkNew.cell(row=lineCnt, column=4).value = 'text'
								sheetmkNew.cell(row=lineCnt, column=5).value = ''
							elif row[1] == 'decimal':
								sheetmkNew.cell(row=lineCnt, column=4).value = row[5].replace('decimal', 'number')
								sheetmkNew.cell(row=lineCnt, column=5).value = ''
							else:
								sheetmkNew.cell(row=lineCnt, column=4).value = row[1]
								sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
							sheetmkNew.cell(row=lineCnt, column=6).value = ''
							lineCnt += 1
						check = False
					if check:
						sheetmkNew.cell(row=lineCnt, column=2).value = row[0]
						sheetmkNew.cell(row=lineCnt, column=3).value = row[3]
						if row[1] == 'varchar':
							sheetmkNew.cell(row=lineCnt, column=4).value = 'string'
							sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
						elif row[1] == 'longtext':
							sheetmkNew.cell(row=lineCnt, column=4).value = 'text'
							sheetmkNew.cell(row=lineCnt, column=5).value = ''
						elif row[1] == 'decimal':
							sheetmkNew.cell(row=lineCnt, column=4).value = row[5].replace('decimal', 'number')
							sheetmkNew.cell(row=lineCnt, column=5).value = ''
						else:
							sheetmkNew.cell(row=lineCnt, column=4).value = row[1]
							sheetmkNew.cell(row=lineCnt, column=5).value = row[2]
						sheetmkNew.cell(row=lineCnt, column=6).value = ''
						lineCnt += 1
				lineCnt += 1
				self.info['Progress'].config(value=count)
				self.info['Percent'].config(text=str(count)+' / '+str(len(self.info['DEListBox'].curselection())) + ' Tables')
				count += 1
			wb.save(self.info['DEPath'])
			self.info['Thread'].stopFunction(False)
			self.info['Status'].join()
			self.info['Window'].destroy()
			self.textB.delete(1.0, END)
			self.textB.insert(1.0, 'DB Scheme -> Excel Complete!\n\n')
			self.textB.insert(END, self.info['DEPath'] + '\n\n')
			f = open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'a')
			f.write(self.datetime.strftime('[ %Y-%m-%d %H:%M:%S ]') + "%-40s" % '\t\tDB Scheme -> Excel File Function.' + "%-69s" % ('[ ' + self.info['DEPath']) + ' ]\n')
			f.close()
			with open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'r') as f:
				lines = f.readlines()
				if len(lines) > 20:
					for i in range(0, len(lines)-20):
						del lines[0]
				for line in lines:
					self.textB.insert(END, line)
			self.textB.config(state=DISABLED)
		except IOError:
			messagebox.showwarning('Warning','Please select a Excel file to save.')
			self.info['Thread'].stopFunction(False)
			self.info['Status'].join()
			self.textB.delete(1.0, END)
		except ValueError:
			messagebox.showwarning('Warning','Please fill out the Sheet name.')
			self.info['Thread'].stopFunction(False)
			self.info['Status'].join()
			self.textB.delete(1.0, END)

	def DB_SQLFunction(self):
		try:
			SQLsentence = str()
			self.info['Cursor'].execute("SELECT TABLE_NAME, TABLE_COMMENT FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA='" + self.info['sid'] + "'")
			tableComment = self.info['Cursor'].fetchall()
			count = 1
			self.info['Progress'].config(maximum=len(self.info['DSListBox'].curselection()))
			for index in self.info['DSListBox'].curselection():
				tableName = str(self.info['DSListBox'].get(index)[0])
				checkNull = False
				self.info['Cursor'].execute("SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, COLUMN_COMMENT, IS_NULLABLE, COLUMN_TYPE FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA='" + self.info['sid'] + "' AND TABLE_NAME='" + tableName + "'")
				sqlList = self.info['Cursor'].fetchall()
				for comment in tableComment:
					if comment[0] == tableName:
						tablecomm = comment[1]
				SQLsentence += 'CREATE TABLE ' + tableName + ' \n(\n'
				rowList = list()
				constraintName = list()
				for sqllist in sqlList:
					rowList.append([str(sqllist[1]), str(sqllist[2]), str(sqllist[3]), str(sqllist[4]), str(sqllist[5]), str(sqllist[6])])
				for row in rowList:
					if row == rowList[-1]:
						if row[4] == 'NO':
							if row[1] == 'int':
								SQLsentence += '\t' + row[0] + ' ' + row[1] + " PRIMARY KEY NOT NULL COMMENT '" + row[3] + "'"
							elif row[1] == 'decimal':
								SQLsentence += '\t' + row[0] + ' ' + row[5].replace('decimal', 'numeric') + " NOT NULL COMMENT '" + row[3] + "'"
								constraintName.append(row[0])
								checkNull = True
							elif row[1] == 'longtext':
								SQLsentence += '\t' + row[0] + " LONGTEXT NOT NULL COMMENT '" + row[3] + "'"
								constraintName.append(row[0])
								checkNull = True
							else:
								SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ") NOT NULL COMMENT '" + row[3] + "'"
								constraintName.append(row[0])
								checkNull = True
						else:
							if row[1] == 'int':
								SQLsentence += '\t' + row[0] + " int PRIMARY KEY NOT NULL COMMENT '" + row[3] + "'"
							elif row[1] == 'decimal':
								SQLsentence += '\t' + row[0] + ' ' + row[5].replace('decimal', 'numeric') + " COMMENT '" + row[3] + "'"
							elif row[1] == 'longtext':
								SQLsentence += '\t' + row[0] + " LONGTEXT COMMENT '" + row[3] + "'"
							else:
								SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ") COMMENT '" + row[3] + "'"
					else:
						if row[4] == 'NO':
							if row[1] == 'int':
								SQLsentence += '\t' + row[0] + ' ' + row[1] + " PRIMARY KEY NOT NULL COMMENT '" + row[3] + "',\n"
							elif row[1] == 'decimal':
								SQLsentence += '\t' + row[0] + ' ' + row[5].replace('decimal', 'numeric') + " NOT NULL COMMENT '" + row[3] + "',\n"
								constraintName.append(row[0])
								checkNull = True
							elif row[1] == 'longtext':
								SQLsentence += '\t' + row[0] + " LONGTEXT NOT NULL COMMENT '" + row[3] + "',\n"
								constraintName.append(row[0])
								checkNull = True
							else:
								SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ") NOT NULL COMMENT '" + row[3] + "',\n"
								constraintName.append(row[0])
								checkNull = True
						else:
							if row[1] == 'int':
								SQLsentence += '\t' + row[0] + " int PRIMARY KEY NOT NULL COMMENT '" + row[3] + "',\n"
							elif row[1] == 'decimal':
								SQLsentence += '\t' + row[0] + ' ' + row[5].replace('decimal', 'numeric') + " COMMENT '" + row[3] + "',\n"
							elif row[1] == 'longtext':
								SQLsentence += '\t' + row[0] + " LONGTEXT COMMENT '" + row[3] + "',\n"
							else:
								SQLsentence += '\t' + row[0] + ' ' + row[1] + '(' + row[2] + ") COMMENT '" + row[3] + "',\n"
				if checkNull:
					constList = ''
					for const in constraintName:
						if const != constraintName[-1]:
							constList += const + ', '
						else:
							constList += const
					SQLsentence += ',\n\n\tCONSTRAINT UK_' + str(self.info['DSListBox'].get(index)[0]).split('_')[1] + ' UNIQUE(' + constList + ')'
				SQLsentence = SQLsentence + "\n) COMMENT '" + tablecomm + "' \nCOLLATE='utf8_bin'\nENGINE=InnoDB\n;\n\n"
				self.info['Progress'].config(value=count)
				self.info['Percent'].config(text=str(count)+' / '+str(len(self.info['DSListBox'].curselection())) + ' Tables')
				count += 1
			f = open(self.info['DSPath'], 'w')
			f.write(SQLsentence)
			f.close()
			self.info['Thread'].stopFunction(False)
			self.info['Status'].join()
			self.info['Window'].destroy()
			self.textB.delete(1.0, END)
			self.textB.insert(1.0, 'DB Scheme -> SQL File Complete!\n\n')
			self.textB.insert(END, self.info['DSPath'] + '\n\n')
			f = open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'a')
			f.write(self.datetime.strftime('[ %Y-%m-%d %H:%M:%S ]') + "%-40s" % '\t\tDB Scheme -> SQL File Function.' + "%-69s" % ('[ ' + self.info['DSPath']) + ' ]\n')
			f.close()
			with open('C:\\Users\\Secuve\\Desktop\\Database Tool\\log\\log.txt', 'r') as f:
				lines = f.readlines()
				if len(lines) > 20:
					for i in range(0, len(lines)-20):
						del lines[0]
				for line in lines:
					self.textB.insert(END, line)
			self.textB.config(state=DISABLED)
		except IOError:
			messagebox.showwarning('Warning','Please select a SQL file to save.')